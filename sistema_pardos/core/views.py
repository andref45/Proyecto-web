from dataclasses import fields
from functools import cache
import json
from tkinter.font import Font
from arrow import now
from django.forms import DurationField, FloatField
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required 
from django.contrib.auth import logout
from django.urls import reverse
from openpyxl import Workbook
from prompt_toolkit import HTML
from sqlalchemy import Cast
from .forms import CustomUserCreationForm
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import ExpressionWrapper
from django.http import JsonResponse
from decimal import Decimal
from django.db.models import Sum, F, Count, Q
from datetime import datetime, timedelta
from .models import MaterialType, Color, Board, Inventory, OrderNotification
from .forms import MaterialTypeForm, ColorForm, BoardForm, InventoryMovementForm
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum, Avg
from .models import ProductionRecord
from .forms import ProductionRecordForm, QuickProductionEntryForm
import csv
from django.http import HttpResponse
from django.template.loader import render_to_string
import tempfile
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Order
from .forms import OrderForm
from django.template.loader import get_template
from openpyxl.styles import PatternFill
from django.db.models.functions import Greatest, Extract
from django.db import transaction

@login_required
def home(request):
    # Estadísticas generales (solo para staff)
    if request.user.is_staff:
        total_products = Board.objects.filter(is_active=True).count()
        products_low_stock = Board.objects.filter(
            is_active=True,
            stock__lte=F('minimum_stock')
        ).count()
        total_value = Board.objects.filter(is_active=True).aggregate(
            total=Sum(F('stock') * F('price_per_m2') * F('width') * F('height'))
        )['total'] or Decimal('0')
    else:
        total_products = 0
        products_low_stock = 0
        total_value = Decimal('0')

    today = timezone.now().date()
    this_month = today.replace(day=1)

    # Estadísticas de órdenes
    if request.user.is_staff:
        orders_query = Order.objects.all()
    else:
        orders_query = Order.objects.filter(customer=request.user)

    orders_stats = {
        'total_today': orders_query.filter(created_at__date=today).count(),
        'total_month': orders_query.filter(created_at__date__gte=this_month).count(),
        'pending': orders_query.filter(status='pending').count(),
        'processing': orders_query.filter(status__in=['processing', 'cutting', 'edge_banding']).count(),
        'completed': orders_query.filter(status__in=['completed', 'delivered']).count(),
        'total_meters_month': orders_query.filter(
            created_at__date__gte=this_month
        ).aggregate(
            total=Sum('total_meters')
        )['total'] or 0
    }

    # Datos para el gráfico (solo para staff)
    if request.user.is_staff:
        last_days = 7
        orders_by_day = Order.objects.filter(
            created_at__date__gte=today - timedelta(days=last_days)
        ).values('created_at__date').annotate(
            count=Count('id'),
            meters=Sum('total_meters')
        ).order_by('created_at__date')

        chart_data = {
            'labels': json.dumps([d['created_at__date'].strftime('%d/%m') for d in orders_by_day]),
            'orders': json.dumps([d['count'] for d in orders_by_day]),
            'meters': json.dumps([float(d['meters'] or 0) for d in orders_by_day])
        }
    else:
        chart_data = {
            'labels': json.dumps([]),
            'orders': json.dumps([]),
            'meters': json.dumps([])
        }

    # Filtros de pedidos
    status_filter = request.GET.get('status', '')
    date_filter = request.GET.get('date', '')

    # Query base para pedidos
    if request.user.is_staff:
        recent_orders = Order.objects.all()
    else:
        recent_orders = Order.objects.filter(customer=request.user)

    # Aplicar filtros
    if status_filter:
        recent_orders = recent_orders.filter(status=status_filter)
    if date_filter:
        recent_orders = recent_orders.filter(created_at__date=date_filter)
    
    recent_orders = recent_orders.order_by('-created_at')[:10]  

    context = {
        'total_products': total_products,
        'products_low_stock': products_low_stock,
        'total_value': total_value,
        'recent_orders': recent_orders,
        'order_status_choices': Order.STATUS_CHOICES,
        'selected_status': status_filter,
        'selected_date': date_filter,
        'orders_stats': orders_stats,
        'chart_data': chart_data,
        'is_staff': request.user.is_staff  
    }
    return render(request, 'core/home.html', context)

@login_required
def products(request):
    return render(request, 'core/products.html')

def exit(request):
    logout(request)
    return redirect('home')

def register(request):
    data = {
        'form' : CustomUserCreationForm()
    }
    if request.method == 'POST':
        user_creation_form = CustomUserCreationForm(data=request.POST)
        if user_creation_form.is_valid():
            user_creation_form.save()

            user = authenticate(username = user_creation_form.cleaned_data['username'], password = user_creation_form.cleaned_data['password1'])
            login(request, user)

            return redirect('home')
    return render(request, 'registration/register.html', data)

# Nuevas vistas para gestión de materiales
@login_required
def material_list(request):
    materials = MaterialType.objects.all()
    return render(request, 'core/materials/list.html', {'materials': materials})

@login_required
def material_add(request):
    if request.method == 'POST':
        form = MaterialTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Material agregado correctamente')
            return redirect('material_list')
    else:
        form = MaterialTypeForm()
    return render(request, 'core/materials/form.html', {'form': form, 'action': 'Agregar'})

@login_required
def material_edit(request, pk):
    material = get_object_or_404(MaterialType, pk=pk)
    if request.method == 'POST':
        form = MaterialTypeForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            messages.success(request, 'Material actualizado correctamente')
            return redirect('material_list')
    else:
        form = MaterialTypeForm(instance=material)
    return render(request, 'core/materials/form.html', {'form': form, 'action': 'Editar'})

# Vistas para gestión de colores
@login_required
def color_list(request):
    colors = Color.objects.all()
    return render(request, 'core/colors/list.html', {'colors': colors})

@login_required
def color_add(request):
    if request.method == 'POST':
        form = ColorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Color agregado correctamente')
            return redirect('color_list')
    else:
        form = ColorForm()
    return render(request, 'core/colors/form.html', {'form': form, 'action': 'Agregar'})

@login_required
def color_edit(request, pk):
    color = get_object_or_404(Color, pk=pk)
    if request.method == 'POST':
        form = ColorForm(request.POST, instance=color)
        if form.is_valid():
            form.save()
            messages.success(request, 'Color actualizado correctamente')
            return redirect('color_list')
    else:
        form = ColorForm(instance=color)
    return render(request, 'core/colors/form.html', {'form': form, 'action': 'Editar'})

# Vistas para gestión de tableros
@login_required
def board_list(request):
    # Obtener filtros
    material_id = request.GET.get('material')
    color_id = request.GET.get('color')
    stock_status = request.GET.get('stock_status')
    
    # Consulta base
    boards = Board.objects.filter(is_active=True)
    
    # Aplicar filtros
    if material_id:
        boards = boards.filter(material_type_id=material_id)
    if color_id:
        boards = boards.filter(color_id=color_id)
    if stock_status == 'low':
        boards = boards.filter(stock__lte=F('minimum_stock'))
    elif stock_status == 'ok':
        boards = boards.filter(stock__gt=F('minimum_stock'))
    
    # Calcular estadísticas
    total_value = sum(board.stock * board.price_per_m2 * board.width * board.height for board in boards)
    low_stock_count = sum(1 for board in boards if board.needs_restock)
    
    # Obtener listas para filtros
    materials = MaterialType.objects.all()
    colors = Color.objects.filter(is_active=True)
    
    context = {
        'boards': boards,
        'materials': materials,
        'colors': colors,
        'total_value': total_value,
        'low_stock_count': low_stock_count,
    }
    
    return render(request, 'core/boards/list.html', context)

@login_required
def board_add(request):
    if request.method == 'POST':
        form = BoardForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tablero agregado correctamente')
            return redirect('board_list')
    else:
        form = BoardForm()
    return render(request, 'core/boards/form.html', {'form': form, 'action': 'Agregar'})

@login_required
def board_edit(request, pk):
    board = get_object_or_404(Board, pk=pk)
    if request.method == 'POST':
        form = BoardForm(request.POST, instance=board)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tablero actualizado correctamente')
            return redirect('board_list')
    else:
        form = BoardForm(instance=board)
    return render(request, 'core/boards/form.html', {'form': form, 'action': 'Editar'})

# Vista para el dashboard con estadísticas
@login_required
def dashboard(request):
    board_stats = Board.objects.filter(is_active=True).aggregate(
        total_products=Count('id'),
        products_low_stock=Count('id', filter=Q(stock__lte=F('minimum_stock'))),
        total_value=Sum(F('stock') * F('price_per_m2') * F('width') * F('height'))
    )

    context = {
        'total_products': board_stats['total_products'],
        'products_low_stock': board_stats['products_low_stock'],
        'total_value': board_stats['total_value'] or Decimal('0')
    }

    return render(request, 'core/dashboard.html', context)


@login_required
def board_delete(request, pk):
    board = get_object_or_404(Board, pk=pk)
    if request.method == 'POST':
        board.is_active = False  
        board.save()
        messages.success(request, 'Tablero eliminado correctamente')
        return redirect('board_list')
        
    return render(request, 'core/boards/delete_confirm.html', {'board': board})



def get_daily_stats():
    today = timezone.now().date()
    records = ProductionRecord.objects.filter(date=today)
    
    return {
        'total_meters': records.aggregate(Sum('meters_cut'))['meters_cut__sum'] or 0,
        'total_pieces': records.aggregate(Sum('pieces_cut'))['pieces_cut__sum'] or 0,
        'total_edges': records.aggregate(Sum('edges_applied'))['edges_applied__sum'] or 0,
        'avg_waste': records.aggregate(Avg('waste_percentage'))['waste_percentage__avg'] or 0,
    }

@login_required
def production_list(request):
    today = timezone.now().date()
    production_records = ProductionRecord.objects.filter(date=today)
    daily_stats = get_daily_stats()
    quick_form = QuickProductionEntryForm()

    context = {
        'production_records': production_records,
        'daily_stats': daily_stats,
        'quick_form': quick_form,
    }
    return render(request, 'production/production_list.html', context)

@login_required
def production_add(request):
    if request.method == 'POST':
        form = ProductionRecordForm(request.POST)
        if form.is_valid():
            production_record = form.save(commit=False)
            production_record.operator = request.user 
            production_record.date = timezone.now().date()
            production_record.save()
            messages.success(request, 'Registro de producción creado exitosamente.')
            return redirect('production_list')
    else:
        form = ProductionRecordForm()  
    
    return render(request, 'production/production_form.html', {'form': form, 'title': 'Nuevo Registro'})

@login_required
def production_edit(request, pk):
    production_record = get_object_or_404(ProductionRecord, pk=pk)
    
    if request.method == 'POST':
        form = ProductionRecordForm(request.POST, instance=production_record)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro de producción actualizado exitosamente.')
            return redirect('production_list')
    else:
        form = ProductionRecordForm(instance=production_record)
    
    return render(request, 'production/production_form.html', {'form': form, 'title': 'Editar Registro'})

@login_required
def production_delete(request, pk):
    production_record = get_object_or_404(ProductionRecord, pk=pk)
    production_record.delete()
    messages.success(request, 'Registro de producción eliminado exitosamente.')
    return redirect('production_list')

@login_required
def quick_production_entry(request):
    if request.method == 'POST':
        form = QuickProductionEntryForm(request.POST)
        if form.is_valid():
            production_record = form.save(commit=False)
            production_record.operator = request.user
            production_record.date = timezone.now().date()
            production_record.start_time = timezone.now().time()
            production_record.end_time = timezone.now().time()
            production_record.save()
            messages.success(request, 'Registro rápido creado exitosamente.')
    return redirect('production_list')



from django.http import HttpResponse
from django.template.loader import render_to_string
from datetime import datetime




@login_required
def order_create(request):
    if request.method == 'POST':
        form = OrderForm(request.POST, request.FILES)
        if form.is_valid():
            order = form.save(commit=False)
            order.customer = request.user
            order._current_user = request.user  # Agregar esta línea
            order.save()
            
            messages.success(request, 'Pedido creado exitosamente.')
            if 'measurements' in request.POST:
                return redirect('order_measurements', pk=order.pk)
            return redirect('order_detail', pk=order.pk)  # Redirigir al detalle del pedido
    else:
        form = OrderForm()
    
    return render(request, 'core/orders/form.html', {
        'form': form,
        'title': 'Nuevo Pedido'
    })

@login_required
def order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)
    
    if not request.user.is_staff and order.customer != request.user:
        messages.error(request, 'No tienes permiso para ver este pedido.')
        return redirect('home')
    
    context = {
        'order': order,
        'order_status_choices': Order.STATUS_CHOICES
    }
    return render(request, 'core/orders/detail.html', context) 



@login_required
def order_update_status(request, pk):
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('home')
        
    order = get_object_or_404(Order, pk=pk)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Order.STATUS_CHOICES):
            old_status = order.get_status_display()
            # Asignar el usuario actual
            order._current_user = request.user
            order.status = new_status
            order.save()
            
            messages.success(
                request, 
                f'Estado del pedido actualizado de "{old_status}" a "{order.get_status_display()}".'
            )
        else:
            messages.error(request, 'Estado inválido.')
            
    return redirect('order_detail', pk=pk)


# En views.py
@login_required
def order_measurements(request, pk):
    order = get_object_or_404(Order.objects.select_related('customer'), pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                data = json.loads(request.body)
                measurements = data.get('measurements', [])
                
                # Validación rápida
                for m in measurements:
                    if not all(0 < m.get(k, 0) <= limit for k, limit in [
                        ('largo', 3.66), ('ancho', 2.44)
                    ]):
                        raise ValueError('Medidas inválidas')
                
                order.measurements = measurements
                order.save()
                
                # Invalidar caché
                cache.delete(f'order_details_{pk}')
                
                return JsonResponse({'success': True})
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
            
    return render(request, 'core/orders/measurements.html', {'order': order})


@login_required
def inventory_dashboard(request):
    # Obtener productos con problemas
    low_stock_products = Board.objects.filter(
        is_active=True,
        stock__lte=F('minimum_stock')
    )
    
    stale_products = Board.objects.filter(
        is_active=True,
        last_movement_date__lte=timezone.now() - timedelta(days=60)
    )
    
    # Estadísticas de rotación por color
    color_stats = Board.objects.values(
        'color__name'
    ).annotate(
        total_products=Count('id'),
        total_stock=Sum('stock'),
        avg_days_without_movement=Avg(
            ExpressionWrapper(
                timezone.now().date() - F('last_movement_date'),
                output_field=DurationField()
            )
        )
    ).order_by('-total_stock')

    context = {
        'low_stock_products': low_stock_products,
        'stale_products': stale_products,
        'color_stats': color_stats,
        'total_products': Board.objects.filter(is_active=True).count(),
        'total_value': Board.objects.filter(is_active=True).aggregate(
            total=Sum(F('stock') * F('price_per_m2') * F('width') * F('height'))
        )['total'] or 0
    }
    
    return render(request, 'inventory/dashboard.html', context)

@login_required
def production_dashboard(request):
    today = timezone.now().date()
    
    if request.method == 'POST':
        try:
            # Obtener y validar las horas del formulario
            start_time = request.POST.get('start_time')
            end_time = request.POST.get('end_time')
            
            if not start_time or not end_time:
                raise ValueError("Las horas de inicio y fin son requeridas")
                
            # Crear el registro usando las horas del formulario
            production_record = ProductionRecord.objects.create(
                operator=request.user,
                date=today,
                start_time=start_time,  # Usar la hora del formulario
                end_time=end_time,      # Usar la hora del formulario
                meters_cut=float(request.POST.get('meters_cut')),
                pieces_cut=int(request.POST.get('pieces_cut')),
                waste_percentage=float(request.POST.get('waste_percentage')),
                edges_applied=request.POST.get('edges_applied')
            )
            messages.success(request, 'Registro guardado exitosamente')
            return redirect('production_list')
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'Error al guardar el registro: {str(e)}')
    
    # Estadísticas del día
    stats = ProductionRecord.objects.filter(
        date=today
    ).aggregate(
        total_meters=Sum('meters_cut'),
        total_pieces=Sum('pieces_cut'),
        total_edges=Sum('edges_applied'),
        avg_waste=Avg('waste_percentage')
    )
    
    # Registros del día
    productions = ProductionRecord.objects.filter(
        date=today
    ).select_related('operator').order_by('-date', '-start_time')
    
    context = {
        'stats': stats,
        'productions': productions,
    }

    if request.GET.get('export') == 'csv':
        return export_production_csv(request)
    
    return render(request, 'production/dashboard.html', context)

login_required
def export_production_csv(request):
    """
    Exporta los registros de producción a CSV asegurando que los datos del operador
    estén correctamente incluidos
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="produccion_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # Configurar el writer con el encoding correcto para caracteres especiales
    response.write(u'\ufeff'.encode('utf-8'))  # BOM para Excel
    writer = csv.writer(response, delimiter=';')  # Usar ; para mejor compatibilidad con Excel
    
    # Escribir encabezados
    writer.writerow([
        'Operador',
        'Fecha',
        'Inicio',
        'Fin',
        'Metros',
        'Piezas',
        'Cantos',
        'Desperdicio (%)'
    ])
    
    # Obtener registros con información del operador
    records = (ProductionRecord.objects
              .select_related('operator')
              .all()
              .order_by('-date', '-start_time'))
    
    # Escribir datos
    for record in records:
        # Obtener nombre del operador de manera segura
        operator_name = ''
        if record.operator:
            if record.operator.first_name or record.operator.last_name:
                operator_name = f"{record.operator.first_name} {record.operator.last_name}".strip()
            else:
                operator_name = record.operator.username
                
        writer.writerow([
            operator_name,
            record.date.strftime('%d/%m/%Y'),
            record.start_time.strftime('%H:%M') if record.start_time else '',
            record.end_time.strftime('%H:%M') if record.end_time else '',
            f"{record.meters_cut:.2f}",  # Sin 'm' para facilitar uso en Excel
            record.pieces_cut,
            f"{record.edges_applied:.2f}" if record.edges_applied else '0.00',
            f"{record.waste_percentage:.1f}"
        ])
    
    return response

# Función auxiliar que podrías usar si necesitas el formato del operador en otros lugares
def get_operator_display_name(operator):
    """
    Obtiene el nombre de visualización del operador de manera consistente
    """
    if not operator:
        return ''
    
    if operator.first_name or operator.last_name:
        return f"{operator.first_name} {operator.last_name}".strip()
    return operator.username

@login_required
def production_add(request):
    if request.method == 'POST':
        try:
            ProductionRecord.objects.create(
                operator=request.user,
                meters_cut=float(request.POST.get('meters_cut')),
                pieces_cut=int(request.POST.get('pieces_cut')),
                waste_percentage=float(request.POST.get('waste_percentage')),
                date=timezone.now().date()
            )
            messages.success(request, 'Registro guardado exitosamente')
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})



@login_required
def production_reports(request):
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    if request.GET.get('start_date'):
        start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
    if request.GET.get('end_date'):
        end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()

    # Obtener registros de producción
    productions = ProductionRecord.objects.filter(
        date__range=[start_date, end_date]
    )

    # Crear datos para el gráfico manualmente
    daily_data = {}
    for prod in productions:
        date_str = prod.date.strftime('%d/%m')
        if date_str not in daily_data:
            daily_data[date_str] = {
                'total_meters': 0,
                'total_pieces': 0
            }
        daily_data[date_str]['total_meters'] += prod.meters_cut
        daily_data[date_str]['total_pieces'] += prod.pieces_cut

    # Convertir a listas ordenadas para el gráfico
    dates = sorted(daily_data.keys())
    meters = [daily_data[date]['total_meters'] for date in dates]

    # Estadísticas generales
    stats = {
        'total_meters': sum(prod.meters_cut for prod in productions),
        'total_pieces': sum(prod.pieces_cut for prod in productions),
        'total_edges': sum(prod.edges_applied or 0 for prod in productions),
        'avg_waste': sum(prod.waste_percentage for prod in productions) / len(productions) if productions else 0
    }

    # Producción por operador
    operator_stats = {}
    for prod in productions:
        op_name = f"{prod.operator.first_name} {prod.operator.last_name}"
        if op_name not in operator_stats:
            operator_stats[op_name] = {
                'total_meters': 0,
                'total_pieces': 0,
                'total_edges': 0,
                'waste_values': []
            }
        operator_stats[op_name]['total_meters'] += prod.meters_cut
        operator_stats[op_name]['total_pieces'] += prod.pieces_cut
        operator_stats[op_name]['total_edges'] += (prod.edges_applied or 0)
        operator_stats[op_name]['waste_values'].append(prod.waste_percentage)

    operator_production = [
        {
            'name': name,
            'total_meters': stats['total_meters'],
            'total_pieces': stats['total_pieces'],
            'total_edges': stats['total_edges'],
            'avg_waste': sum(stats['waste_values']) / len(stats['waste_values']) if stats['waste_values'] else 0
        }
        for name, stats in operator_stats.items()
    ]

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'stats': stats,
        'operator_production': operator_production,
        'chart_dates': json.dumps(dates),
        'chart_meters': json.dumps(meters)
    }

    return render(request, 'reports/production.html', context)

def export_production_excel(request, context):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Producción"

    # Encabezados
    headers = ['Fecha', 'Operador', 'Metros Cortados', 'Piezas', 'Cantos', 'Desperdicio (%)']
    ws.append(headers)

    # Datos
    productions = ProductionRecord.objects.filter(
        date__range=[context['start_date'], context['end_date']]
    ).select_related('operator')

    for prod in productions:
        ws.append([
            prod.date,
            prod.operator.get_full_name(),
            prod.meters_cut,
            prod.pieces_cut,
            prod.edges_applied or 0,
            prod.waste_percentage
        ])

    # Estilo
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=produccion.xlsx'
    
    wb.save(response)
    return response


@login_required
def board_rotation_report(request):
    """Vista para el reporte de rotación de productos"""
    boards = Board.objects.filter(is_active=True).annotate(
        days_inactive=timezone.now().date() - F('last_movement_date')
    ).order_by('-days_inactive')

    context = {
        'boards': boards,
        'no_movement_60': boards.filter(days_inactive__gte=60).count(),
        'no_movement_30': boards.filter(days_inactive__gte=30, days_inactive__lt=60).count(),
        'no_movement_15': boards.filter(days_inactive__gte=15, days_inactive__lt=30).count(),
    }
    return render(request, 'reports/board_rotation.html', context)

@login_required
def production_efficiency(request):
    """Vista para el análisis de eficiencia de producción"""
    today = timezone.now().date()
    start_date = today - timedelta(days=30)
    
    records = ProductionRecord.objects.filter(
        date__range=[start_date, today]
    ).select_related('operator')

    # Análisis por operador
    operator_stats = records.values(
        'operator__username'
    ).annotate(
        total_hours=Sum(
            ExpressionWrapper(
                F('end_time') - F('start_time'),
                output_field=DurationField()
            )
        ),
        total_meters=Sum('meters_cut'),
        total_pieces=Sum('pieces_cut'),
        avg_waste=Avg('waste_percentage')
    )

    context = {
        'operator_stats': operator_stats,
        'start_date': start_date,
        'end_date': today,
    }
    return render(request, 'production/efficiency.html', context)

@login_required
def material_movement_history(request, board_id):
    """Vista para el historial de movimientos de un material"""
    board = get_object_or_404(Board, id=board_id)
    movements = Inventory.objects.filter(
        board=board
    ).order_by('-date')

    context = {
        'board': board,
        'movements': movements,
    }
    return render(request, 'inventory/movement_history.html', context)

@login_required
def low_stock_alert(request):
    """Vista para alertas de stock bajo"""
    low_stock_boards = Board.objects.filter(
        is_active=True,
        stock__lte=F('minimum_stock')
    ).order_by('stock')

    context = {
        'boards': low_stock_boards,
        'total_alerts': low_stock_boards.count()
    }
    return render(request, 'inventory/low_stock_alert.html', context)


@login_required
def realtime_dashboard(request):
    # Obtener datos actuales
    current_data = {
        'production': ProductionRecord.objects.filter(
            date=timezone.now().date()
        ).aggregate(
            total_meters=Sum('meters_cut'),
            total_pieces=Sum('pieces_cut'),
            total_edges=Sum('edges_applied'),
            avg_waste=Avg('waste_percentage')
        ),
        'operators': ProductionRecord.objects.filter(
            date=timezone.now().date()
        ).values(
            'operator__username',
            'operator__first_name'
        ).annotate(
            total_meters=Sum('meters_cut'),
            total_pieces=Sum('pieces_cut'),
            efficiency=ExpressionWrapper(
                Sum('meters_cut') * 100.0 / Cast('pieces_cut', FloatField()),
                output_field=FloatField()
            )
        ),
        'alerts': StockAlert.objects.filter(is_active=True)[:5]
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse(current_data)
    
    return render(request, 'dashboard/realtime.html', {
        'initial_data': current_data
    })


from django.db.models import Sum, Avg, Count, F
from .models import StockAlert

@login_required
def dashboard_realtime(request):
    today = timezone.now().date()
    
    # Usar select_related para reducir queries
    recent_orders = Order.objects.select_related('customer').filter(
        created_at__date=today
    ).order_by('-created_at')[:5]

    context = {
        'orders': recent_orders,
        'stats': {
            'pending_orders': Order.objects.filter(status='pending').count(),
            'active_orders': Order.objects.filter(
                status__in=['processing', 'cutting']
            ).count(),
            'completed_today': Order.objects.filter(
                status='completed',
                updated_at__date=today
            ).count()
        },
        'alerts': StockAlert.objects.filter(is_active=True)[:5]
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse(context)
    
    return render(request, 'core/dashboard.html', context)

@login_required
def update_alerts(request):
    """Vista para generar y actualizar alertas"""
    StockAlert.create_alerts()  # Método que creamos en el modelo
    return JsonResponse({'success': True})

# En views.py, agrega esta nueva función
@login_required
def update_dashboard_data(request):
    """Endpoint para actualizar datos del dashboard vía AJAX"""
    current_time = timezone.now()
    
    production_data = ProductionRecord.objects.filter(
        date=current_time.date()
    ).aggregate(
        total_meters=Sum('meters_cut'),
        total_pieces=Sum('pieces_cut'),
        avg_waste=Avg('waste_percentage')
    )
    
    # Actualizar alertas
    StockAlert.create_alerts()
    
    return JsonResponse({
        'production': production_data,
        'alerts': list(StockAlert.objects.filter(
            is_active=True
        ).values('message', 'alert_type', 'created_at')[:5])
    })



@login_required
def operator_metrics(request):
    """Vista para mostrar métricas de operadores"""
    # Obtener el rango de fechas seleccionado
    time_range = request.GET.get('range', 'day')
    today = timezone.now().date()
    
    # Determinar fecha inicial según el rango
    if time_range == 'week':
        start_date = today - timedelta(days=7)
    elif time_range == 'month':
        start_date = today - timedelta(days=30)
    else:  # day
        start_date = today

    # Obtener estadísticas de operadores
    operators = ProductionRecord.objects.filter(
        date__range=[start_date, today]
    ).values(
        'operator__username',
        'operator__first_name'
    ).annotate(
        total_meters=Sum('meters_cut'),
        total_pieces=Sum('pieces_cut'),
        avg_efficiency=100 - Avg('waste_percentage'),
        pieces_per_hour=Cast(Sum('pieces_cut'), FloatField()) / 
                       Greatest(Sum(
                           Extract('end_time', 'hour') - Extract('start_time', 'hour')
                       ), 1)
    )

    # Preparar datos para los gráficos
    dates = []
    production_data = []
    current_date = start_date
    
    while current_date <= today:
        dates.append(current_date.strftime('%d/%m'))
        daily_production = ProductionRecord.objects.filter(
            date=current_date
        ).values(
            'operator__username'
        ).annotate(
            total_meters=Sum('meters_cut')
        )
        
        production_data.append({
            'date': current_date.strftime('%d/%m'),
            'data': {p['operator__username']: p['total_meters'] for p in daily_production}
        })
        current_date += timedelta(days=1)

    # Preparar datasets para el gráfico de producción
    operator_names = list(set(op['operator__username'] for op in operators))
    production_datasets = []
    
    for operator in operator_names:
        dataset = {
            'label': operator,
            'data': [p['data'].get(operator, 0) for p in production_data],
            'borderColor': f'hsl({hash(operator) % 360}, 70%, 50%)',
            'fill': False
        }
        production_datasets.append(dataset)

    # Preparar datos de eficiencia
    efficiency_data = [op['avg_efficiency'] for op in operators]

    # Si es una solicitud AJAX, devolver JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'dates': dates,
            'production_datasets': production_datasets,
            'operator_names': operator_names,
            'efficiency_data': efficiency_data
        })

    # Para la carga inicial de la página, pasar contexto al template
    context = {
        'operators': operators,
        'dates': json.dumps(dates),
        'production_datasets': json.dumps(production_datasets),
        'operator_names': json.dumps(operator_names),
        'efficiency_data': json.dumps(efficiency_data)
    }
    
    return render(request, 'core/operator_metrics.html', context)



@login_required
def dashboard_data(request):
    cache_key = f'dashboard_data_{request.user.id}'
    data = cache.get(cache_key)
    
    if not data:
        today = now().date()
        
        # Consulta optimizada para estadísticas
        stats = Order.objects.filter(
            created_at__date=today
        ).aggregate(
            total_today=Count('id'),
            total_meters=Sum('total_meters'),
            pending=Count('id', filter=Q(status='pending')),
            processing=Count('id', filter=Q(status__in=['processing', 'cutting'])),
            completed=Count('id', filter=Q(status='completed'))
        )
        
        data = {
            'stats': stats,
            'last_update': now().isoformat()
        }
        
        cache.set(cache_key, data, 60)  # Cache por 1 minuto
    
    return JsonResponse(data)

@login_required
def get_notifications(request):
    notifications = OrderNotification.objects.select_related('order').filter(
        order__customer=request.user,
        is_read=False
    ).order_by('-created_at')[:5]
    
    return JsonResponse({
        'notifications': [{
            'id': n.id,
            'message': n.message,
            'created_at': n.created_at.strftime('%Y-%m-%d %H:%M'),
            'order_id': n.order_id
        } for n in notifications]
    })

@login_required
def mark_notification_read(request, notification_id):
    OrderNotification.objects.filter(id=notification_id).update(is_read=True)
    return JsonResponse({'success': True})

@login_required
def customer_orders(request):
    """Vista para que los clientes vean sus pedidos"""
    orders = Order.objects.filter(customer=request.user)
    
    status = request.GET.get('status')
    if status:
        orders = orders.filter(status=status)
        
    orders = orders.order_by('-created_at')
    
    context = {
        'orders': orders,
        'status_choices': Order.STATUS_CHOICES,
        'current_status': status
    }
    return render(request, 'orders/customer_orders.html', context)

@login_required
def customer_order_detail(request, pk):
    """Vista detallada del pedido para el cliente"""
    order = get_object_or_404(Order, pk=pk)
    
    if order.customer != request.user and not request.user.is_staff:
        messages.error(request, 'No tienes permiso para ver este pedido.')
        return redirect('customer_orders')
    
    status_history = order.status_history.all().order_by('-created_at')
    
    context = {
        'order': order,
        'status_history': status_history
    }
    return render(request, 'orders/customer_order_detail.html', context)